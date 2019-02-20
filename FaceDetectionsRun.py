import cv2
import os
import numpy as np
from PIL import ImageTk
from PIL import Image as PilImage
import pickle
import sys
import pandas as pd
import datetime
import time
import openpyxl
import datetime
from tkinter import *
import tkinter as tk
from tkinter import messagebox
from tkinter import Message ,Text
from tkinter import simpledialog
import sys


now=datetime.datetime.now()
time=now.time
today=now.day
month=now.month
cap = cv2.VideoCapture(0)
Detect = FALSE

def Detect ():
    face_cascade = cv2.CascadeClassifier('cascades\data\haarcascade_frontalface_alt2.xml')
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read("dataSet/recognizer/trainner.yml")
    labels = {"Name": 1}
    with open("labels.pickle","rb") as f:
        og_labels = pickle.load(f)
        labels = {v:k for k,v in og_labels.items()}
    #timeNum=200; #Runingtime
    while(True):
        #timeNum=timeNum-1;
        #print (timeNum)
        # Capture frame-by-frame
        ret,image=cap.read();
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        faces=face_cascade.detectMultiScale(gray,1.3,5);
        for(x,y,w,h) in faces:
            #print(x,y,w,h)
            roi_gray = gray[y:y+h,x:x+w] #(ycord1,ycord_end)
            roi_color = image[y:y+h,x:x+w]
            #reconizer
            id_,conf = recognizer.predict(roi_gray)
            if conf>=50 and conf <=90:                   
                font = cv2.FONT_HERSHEY_SIMPLEX
                name = labels[id_]
                color = (255,255,255)
                stroke = 2
                cv2.putText(image, name, (x,y),font,1,color,stroke,cv2.LINE_AA)
            img_item = "my-img.jpg"
            cv2.imwrite(img_item,color)
            color = (255,0,0)
            stroke = 2
            end_cord_x = x+w
            end_cord_Y = y+h
            cv2.rectangle(image,(x,y),(end_cord_x,end_cord_Y),color,stroke)
            #print(id_)#เลขลำดับโฟเดอร์
            print(labels[id_])#ปริ้นชื่อโฟเดอร์
            print(datetime.datetime.now())#printtime

            #exprot
            op = openpyxl.load_workbook('Timestamp/Timestamp.xlsx')
            ws = op.active.title
            sh = op["Timestamp"]
            rows=sh.max_row
            cols=sh.min_column
            for i in range(rows,rows+1):
                sh['A'+str(i+1)]=datetime.datetime.now()
                sh['B'+str(i+1)]=labels[id_]
                c=sh.cell(rows+1,1)
            op.save('Timestamp/Timestamp.xlsx')
            cv2.waitKey(1000)
            #print(str(today))
            #print (ws)            
        cv2.imshow("Face",image)
        if cv2.waitKey(1) == ord('q'):
            break;
    message.configure(text="Detection had been stop.") 
    cap.release()
    cv2.destroyAllWindows()       
    
                
def Registion():
    id=input = simpledialog.askstring("input string","Enter your name.")
    message.configure(text=id+"Registion success") 
    faceDetect=cv2.CascadeClassifier('cascades/data/haarcascade_frontalface_alt2.xml');
    cap=cv2.VideoCapture(0);
    recognizer = cv2.face.LBPHFaceRecognizer_create()
    recognizer.read("dataSet/recognizer/trainner.yml")
    idNum=0;
    def createFolder(directory):
        try:
            if not os.path.exists(directory):
                os.makedirs(directory)
        except OSError:
            print ('Error: Creating directory. ' +  directory)
                
    while(True):
        createFolder('./dataSET/user/'+str(id)+'/')
        ret,image=cap.read();
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        faces=faceDetect.detectMultiScale(gray,1.3,5);
        for(x,y,w,h) in faces:
            roi_gray = gray[y:y+h,x:x+w] #(ycord1,ycord_end)
            roi_color = image[y:y+h,x:x+w]
            id_,conf = recognizer.predict(roi_gray)
            if conf>=50 and conf <=90:
                idNum=idNum+1;
                cv2.imwrite("dataSet/User/"+str(id)+"/"+str(id)+"."+str(idNum)+".jpg",gray[y:y+h,x:x+w]);
                cv2.rectangle(image,(x,y),(x+w,y+h),(0,255,0),2)
                cv2.waitKey(100);
        cv2.imshow("Face",image);
        cv2.waitKey(1);
        
        if(idNum>29):
            print("Registration Success.")
            break

    cap.release()
    cv2.destroyAllWindows()

    #Recognizer
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    image_dir = os.path.join(BASE_DIR,"dataSET/User")

    faceDetect = cv2.CascadeClassifier('cascades/data/haarcascade_frontalface_alt2.xml');
    recognizer = cv2.face.LBPHFaceRecognizer_create()

    current_id = 0
    label_ids = {}
    y_labels = []
    x_train = []
    for root,dirs, files in os.walk(image_dir):
        for file in files:
            if file.endswith("jpg"):
                path = os.path.join(root, file)
                label = os.path.basename(root).replace("","").lower()
                #print(label,path)
                if not label in label_ids:
                    label_ids[label]=current_id
                    current_id += 1
                id_ =label_ids[label]
                #print(label_ids)
                #y_labels.append(label)
                #x_train.append(path)
                pil_image = PilImage.open(path).convert("L")
                image_array = np.array(pil_image,"uint8")
                #print(image_array)
                faces = faceDetect.detectMultiScale(image_array,1.3,5);

                for(x,y,w,h) in faces:
                    roi = image_array[y:y+h,x:x+w]
                    x_train.append(roi)
                    y_labels.append(id_)
    #print(y_labels)
    #print(x_train)
    with open("labels.pickle",'wb') as f:
        pickle.dump(label_ids, f)
    recognizer.train(x_train, np.array(y_labels))
    recognizer.save("dataSet/recognizer/trainner.yml")
        
# When everything done, release the capture


def Exit():
    status=messagebox.askyesno(title='Exit',message="Do youwant to exit system?")
    if status>0:
        gui.destroy()
        cap.release()
        exit()
gui = Tk()
gui.geometry("600x275")
gui.configure(bg="darkorange")
gui.title("Face Detection")
headlabel1 = Label(text="Welcom to Facedection System",fg="black",bg="darkorange",width=25,height=1,font=('times', 30, 'italic bold underline'))
headlabel2 = Label(text="",fg="black",bg="darkorange",width=5,height=2,font=('times', 5, 'italic bold'))
Detect_Photo = PhotoImage(file='dataSet/Icon/Detect.png')
StopDetect_Photo = PhotoImage(file='dataSet/Icon/StopDetect.png')
Register_Photo = PhotoImage(file='dataSet/Icon/Register.png')
Exit_Photo = PhotoImage(file='dataSet/Icon/Exit.png')
Detect_b = Button(gui,image = Detect_Photo,bg="darkorange",command=Detect)
Register_b = Button(gui,image = Register_Photo,bg="darkorange",command=Registion)
Exit_b = Button(gui,image = Exit_Photo,bg="darkorange",command=Exit)

Detectb_label= Label(text="Detect",fg="black",bg="darkorange",width=10,height=1,font=('times', 10, 'italic bold'))
Registerb_label= Label(text="Registion",fg="black",bg="darkorange",width=10,height=1,font=('times', 10, 'italic bold'))
Exitb_label= Label(text="Exit",fg="black",bg="darkorange",width=10,height=1,font=('times', 10, 'italic bold'))
headlabel3 = Label(text="",fg="black",bg="darkorange",width=5,height=2,font=('times', 2, 'italic bold'))
message = Label(text="",fg="black",bg="orange",width=25,height=1,font=('times', 20, 'italic bold underline'))

#GRID
headlabel1.grid(row=0,column=1,columnspan=3)
headlabel2.grid(row=1,column=1,columnspan=3)
Detect_b.grid(row=2,column=0,columnspan=3)
Register_b.grid(row=2,column=1,columnspan=3)
Exit_b.grid(row=2,column=2,columnspan=3)
Detectb_label.grid(row=3,column=0,columnspan=3)
Registerb_label.grid(row=3,column=1,columnspan=3)
Exitb_label.grid(row=3,column=2,columnspan=3)
headlabel3.grid(row=5,column=1,columnspan=3)
message.grid(row=6,column=1,columnspan=3)


def tick(time1=''):
    R_time = datetime.datetime.now().strftime("%y-%m-%d %H:%M:%S")
    clock_frame.config(text=R_time)
    clock_frame.after(200,tick)
    
clock_frame = tk.Label(gui,font=('times',20,'bold'),fg="black",bg='darkorange')
clock_frame.grid(row=4,column=1,columnspan=3)
tick()

gui.mainloop()
cap.release()
cv2.destroyAllWindows()
